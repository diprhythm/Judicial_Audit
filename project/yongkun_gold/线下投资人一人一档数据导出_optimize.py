#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
一人一表 · 隐身索引版（MySQL 8.0.23+）
- 为 IMK/姓名建立 INVISIBLE 生成列 + 索引：查询全走索引，且不污染表结构展示
- 懒创建工作表：台账/租赁/流水仅当有数据时才建表
- 并发拉满（≤61）：默认 min(61, max(16, CPU*2))
- IMK 匹配忽略字面 'null'（大小写不敏感）

运行建议：
  首次：python script.py  （会创建/检查隐身列与索引，可能花时间）
  之后：python script.py --skip-index-check  （跳过自检，直接起飞）
选项：
  --no-tx-index        首次先不对流水表建隐身索引（默认会建）
  --tx-remove-null     对流水姓名标准化时也去掉 'NULL' 子串（默认不去）
"""

import os
import re
import sys
import time
import argparse
import multiprocessing as mp
from typing import Tuple, List, Iterable

import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import BoundedSemaphore
from random import random

# ======== MySQL 连接信息 ========
DB_HOST = "localhost"
DB_PORT = 3306
DB_USER = "root"
DB_PASS = "010203"
DB_NAME = "yougkun_gold"

# ======== 表名 ========
T_BASE = "台账匹配流水结果"
T_LEDGER_DEP = "1代保管业务台账明细"
T_LEDGER_LEASE = "2租赁业务台账明细"
T_TX = "永坤资金池账户交易明细_身份映射"

# 默认银行流水匹配字段（可用 --tx-extra-fields 再追加）
DEFAULT_TX_NAME_FIELDS = [
    "对手户名",
    "对手户名_审计专用",
    "对手户名_审计专用_norm",
    "对手户名_审计专用_norm2",
    "映射_户名",
    "映射_户名_norm",
]

# =============== 通用工具 ===============
def make_engine(pool_size: int, max_overflow: int) -> Engine:
    url = f"mysql+pymysql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4"
    return create_engine(
        url,
        pool_pre_ping=True, pool_recycle=3600,
        pool_size=pool_size, max_overflow=max_overflow,
        isolation_level="READ UNCOMMITTED",
        pool_timeout=60, future=True,
    )

def get_mysql_version(engine: Engine) -> Tuple[int, int, int]:
    with engine.connect() as conn:
        ver = conn.execute(text("SELECT VERSION()")).scalar() or "8.0.0"
    m = re.match(r"(\d+)\.(\d+)\.(\d+)", ver)
    return (8,0,0) if not m else tuple(map(int, m.groups()))

def norm_sql_expr(col: str, use_regex: bool) -> str:
    """UPPER + 去空白 + 去常见标点 + 括号半角化"""
    base = f"COALESCE({col}, '')"
    s1 = f"REPLACE(REPLACE({base}, '（', '('), '）', ')')"
    if use_regex:
        s2 = f"REGEXP_REPLACE({s1}, '[[:space:]]+', '')"
        s3 = ("REGEXP_REPLACE("
              f"{s2},"
              " '[[:punct:]，。；：、！？“”‘’—－·•（）()]+'"
              ", '' )")
        return f"UPPER({s3})"
    else:
        s2 = f"REPLACE(REPLACE({s1}, ' ', ''), '　', '')"
        for ch in [',','.', '，','。','；','：','、','！','（','）','(',')','·','•','-','_',';',':']:
            s2 = f"REPLACE({s2}, '{ch}', '')"
        return f"UPPER({s2})"

def norm_imk_expr(col: str, use_regex: bool) -> str:
    """IMK 专用：在通用标准化基础上 REPLACE('NULL','')"""
    base = norm_sql_expr(col, use_regex)
    return f"REPLACE({base}, 'NULL', '')"

def norm_name_expr(col: str, use_regex: bool, remove_null: bool=False) -> str:
    """姓名标准化：默认不去 'NULL'，可选去"""
    base = norm_sql_expr(col, use_regex)
    return f"REPLACE({base}, 'NULL', '')" if remove_null else base

def jitter_sleep(base: float, factor: float, attempt: int):
    time.sleep(base * (factor ** attempt) * (0.7 + 0.6 * random()))

def safe_filename(name: str) -> str:
    name = str(name)
    trans = str.maketrans({c: "_" for c in r'\/:*?"<>|'})
    return name.translate(trans)[:150]

# =============== 隐身生成列 + 索引 ===============
def ensure_invisible_gencol_index(
    engine: Engine,
    table: str,
    source_col: str,
    gen_col: str,
    idx_name: str,
    expr_sql: str
):
    """
    在 `table` 上创建 INVISIBLE 生成列 `gen_col`（表达式 expr_sql）及其索引 `idx_name`。
    已存在则跳过，不重复创建。
    """
    with engine.begin() as conn:
        col_exist = conn.execute(text("""
            SELECT 1 FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA=:db AND TABLE_NAME=:tb AND COLUMN_NAME=:col
        """), {"db": DB_NAME, "tb": table, "col": gen_col}).first() is not None

        if not col_exist:
            # INVISIBLE 仅 8.0.23+；本脚本已在前置检查版本
            ddl = f"""
            ALTER TABLE `{table}`
              ADD COLUMN `{gen_col}` VARCHAR(255)
              GENERATED ALWAYS AS ({expr_sql}) STORED INVISIBLE
            """
            conn.execute(text(ddl))

        idx_exist = conn.execute(text("""
            SELECT 1 FROM information_schema.STATISTICS
            WHERE TABLE_SCHEMA=:db AND TABLE_NAME=:tb AND INDEX_NAME=:idx
        """), {"db": DB_NAME, "tb": table, "idx": idx_name}).first() is not None

        if not idx_exist:
            conn.execute(text(f"CREATE INDEX `{idx_name}` ON `{table}` (`{gen_col}`)"))

# =============== 懒创建写表 ===============
def write_stream_sheet_lazy(wb: Workbook, title: str, cols: List[str], frames: Iterable[pd.DataFrame]) -> bool:
    ws = None
    for ch in frames:
        if ch is None or ch.empty:
            continue
        if ws is None:
            ws = wb.create_sheet(title)
            ws.append(cols)
        local_cols = [c for c in cols if c in ch.columns]
        if len(local_cols) != len(cols):
            for row in ch.itertuples(index=False, name=None):
                row_map = dict(zip(ch.columns, row))
                ws.append([row_map.get(c) for c in cols])
        else:
            for row in ch[local_cols].itertuples(index=False, name=None):
                ws.append(list(row))
    return ws is not None

# =============== 主流程 ===============
def main():
    cpu = mp.cpu_count() or 4
    ap = argparse.ArgumentParser(description="逐项目导出 · 隐身索引版")
    ap.add_argument("--workers", type=int, default=None, help="总并发（默认=min(61, max(16, CPU*2)))")
    ap.add_argument("--tx-workers", type=int, default=None, help="银行流水并发（默认=min(24, max(6, workers//3)))")
    ap.add_argument("--chunk", type=int, default=120_000, help="分块大小（默认 120000）")
    ap.add_argument("--tx-extra-fields", type=str, default="", help="流水表额外姓名列，逗号分隔")
    ap.add_argument("--print-every", type=int, default=5, help="进度打印频率（默认 5）")
    ap.add_argument("--skip-index-check", action="store_true", help="跳过隐身列/索引自检（后续跑建议开启）")
    ap.add_argument("--no-tx-index", action="store_true", help="不对流水表建立隐身列索引")
    ap.add_argument("--tx-remove-null", action="store_true", help="姓名标准化时也移除 'NULL' 子串（默认不移除）")
    args = ap.parse_args()

    out_dir = input("请输入输出文件夹路径：").strip().strip('"').strip("'")
    if not out_dir:
        print("未提供输出路径，已退出。"); sys.exit(1)
    os.makedirs(out_dir, exist_ok=True)
    print(f"[INFO] 导出目录：{out_dir}")

    max_workers = args.workers if args.workers is not None else min(61, max(16, cpu * 2))
    max_workers = min(61, max(1, max_workers))
    max_tx_workers = args.tx_workers if args.tx_workers is not None else min(24, max(6, max_workers // 3))
    max_tx_workers = min(61, max(1, max_tx_workers))
    chunk_size = max(20_000, args.chunk)
    print(f"[INFO] 并发：{max_workers}，TX限流：{max_tx_workers}，chunksize={chunk_size}")

    engine = make_engine(pool_size=max_workers, max_overflow=max_workers)
    major, minor, patch = get_mysql_version(engine)
    if (major, minor, patch) < (8, 0, 23):
        print(f"[FATAL] 需要 MySQL ≥ 8.0.23（当前 {major}.{minor}.{patch}）。"); sys.exit(2)
    use_regex = True
    print(f"[INFO] MySQL {major}.{minor}.{patch}（INVISIBLE/REGEXP 可用）")

    # 银行流水匹配列
    extra = [c.strip() for c in args.tx_extra_fields.split(",") if c.strip()]
    tx_fields: List[str] = list(dict.fromkeys(DEFAULT_TX_NAME_FIELDS + extra))
    print("[INFO] 银行流水匹配列：", ", ".join(tx_fields))

    # === 隐身列/索引 自检创建 ===
    if not args.skip_index_check:
        print("[INFO] 隐身索引自检…（首次可能较久；后续用 --skip-index-check）")
        # 1) 三张台账/汇总：IMK
        imk_expr_base   = norm_imk_expr("`Identity_Matching_Key`", use_regex)
        for tb, idx in [(T_LEDGER_DEP, "idx_dep_imk_norm"),
                        (T_LEDGER_LEASE, "idx_lease_imk_norm"),
                        (T_BASE, "idx_base_imk_norm")]:
            ensure_invisible_gencol_index(
                engine, tb, "Identity_Matching_Key",
                "imk_norm_nulldrop", idx, imk_expr_base
            )

        # 2) 流水表：姓名字段（可关闭）
        if not args.no_tx_index:
            for i, col in enumerate(tx_fields, start=1):
                # 生成一个短小安全的隐身列名/索引名
                gen_col = f"txnm_{i:02d}_norm"   # INVISIBLE 列名
                idx_name = f"idx_tx_nm_{i:02d}"
                expr = norm_name_expr(f"`{col}`", use_regex, remove_null=args.tx_remove_null)
                ensure_invisible_gencol_index(
                    engine, T_TX, col, gen_col, idx_name, expr
                )
        print("[INFO] 隐身索引检查完成。")

    # === 读基础键 & 列头 ===
    with engine.connect() as conn:
        base_keys = pd.read_sql(text(f"""
            SELECT `序号_Primary_Key`, `Identity_Matching_Key`, `投资人姓名`
            FROM `{T_BASE}`
            WHERE COALESCE(TRIM(`Identity_Matching_Key`),'') <> ''
              AND COALESCE(TRIM(`投资人姓名`),'') <> '';
        """), conn)

        cols_base  = [m["Field"] for m in conn.execute(text(f"SHOW COLUMNS FROM `{T_BASE}`")).mappings()]
        cols_dep   = [m["Field"] for m in conn.execute(text(f"SHOW COLUMNS FROM `{T_LEDGER_DEP}`")).mappings()]
        cols_lease = [m["Field"] for m in conn.execute(text(f"SHOW COLUMNS FROM `{T_LEDGER_LEASE}`")).mappings()]
        cols_tx    = [m["Field"] for m in conn.execute(text(f"SHOW COLUMNS FROM `{T_TX}`")).mappings()]

    if base_keys.empty:
        print("[WARN] 基础表无有效键，退出。"); return

    base_keys["__wb_id"] = base_keys["序号_Primary_Key"].astype(str) + "、" + base_keys["Identity_Matching_Key"].astype(str)
    projects = base_keys[["__wb_id","序号_Primary_Key","Identity_Matching_Key","投资人姓名"]].drop_duplicates()
    total = len(projects)
    print(f"[INFO] 需要生成工作簿数量：{total}")

    # === 限流 ===
    tx_sema = BoundedSemaphore(value=max_tx_workers)

    # === SQL（全部索引可用：左侧为隐身生成列；右侧为常量表达式）===
    norm_nm  = lambda col: norm_sql_expr(col, use_regex)
    norm_imk = lambda col: norm_imk_expr(col, use_regex)

    sql_summary = text(f"""
        SELECT *
        FROM `{T_BASE}`
        WHERE `序号_Primary_Key` = :pk
          AND `imk_norm_nulldrop` = {norm_imk(":imk")};
    """)
    sql_dep = text(f"""
        SELECT *
        FROM `{T_LEDGER_DEP}`
        WHERE `imk_norm_nulldrop` = {norm_imk(":imk")}
          AND {norm_imk(":imk")} <> '';
    """)
    sql_lease = text(f"""
        SELECT *
        FROM `{T_LEDGER_LEASE}`
        WHERE `imk_norm_nulldrop` = {norm_imk(":imk")}
          AND {norm_imk(":imk")} <> '';
    """)

    # 银行流水：用隐身列命中索引；把原先的函数比较换成等值比较
    # 我们假设隐身列名为 txnm_{i:02d}_norm （与上面的 ensure 保持一致）
    def build_sql_tx_union(n_fields: int) -> str:
        parts = []
        for i in range(1, n_fields+1):
            gen_col = f"`txnm_{i:02d}_norm`"
            parts.append(f"SELECT t.* FROM `{T_TX}` t WHERE {gen_col} = {norm_nm(':nm')} AND {norm_nm(':nm')} <> ''")
        return "\nUNION\n".join(parts) if parts else "SELECT t.* FROM `{T_TX}` t WHERE 1=0"

    use_tx_index = (not args.no_tx_index)
    if use_tx_index:
        unions = build_sql_tx_union(len(tx_fields))
        sql_tx = text(f"SELECT * FROM (\n{unions}\n) z")
    else:
        # 退回旧逻辑（不建议）：函数在列左侧 → 不可索引
        nm_exprs_col = [norm_nm(f"t.`{c}`") for c in tx_fields]
        nm_expr_param = norm_nm(":nm")
        unions = "\nUNION\n".join(
            f"SELECT t.* FROM `{T_TX}` t WHERE {e} = {nm_expr_param} AND {nm_expr_param} <> ''"
            for e in nm_exprs_col
        )
        sql_tx = text(f"SELECT * FROM (\n{unions}\n) z")

    # === 处理单个项目 ===
    def process_one(row) -> Tuple[bool, str, str, str]:
        wb_id = row["__wb_id"]
        base_pk = row["序号_Primary_Key"]
        imk = (row["Identity_Matching_Key"] or "")
        name = (row["投资人姓名"] or "").strip()

        fn = safe_filename(wb_id) + ".xlsx"
        path = os.path.join(out_dir, fn)

        for attempt in range(5):
            t0 = time.time()
            try:
                with engine.connect() as conn:
                    wb = Workbook(write_only=True)

                    # 汇总（总是建一张，便于核对）
                    ws = wb.create_sheet("汇总表")
                    df_sum = pd.read_sql(sql_summary, conn, params={"pk": base_pk, "imk": imk})
                    if df_sum.empty:
                        ws.append(cols_base)
                    else:
                        ws.append(list(df_sum.columns))
                        for tup in df_sum.itertuples(index=False, name=None):
                            ws.append(list(tup))

                    # 代保管/租赁（懒创建）
                    _dep_created = write_stream_sheet_lazy(
                        wb, "代保管业务明细", cols_dep,
                        pd.read_sql(sql_dep, conn, params={"imk": imk}, chunksize=chunk_size)
                    )
                    _lease_created = write_stream_sheet_lazy(
                        wb, "租赁业务明细", cols_lease,
                        pd.read_sql(sql_lease, conn, params={"imk": imk}, chunksize=chunk_size)
                    )

                    # 银行流水（懒创建 + 限流）
                    if name:
                        tx_sema.acquire()
                        try:
                            _ = write_stream_sheet_lazy(
                                wb, "银行流水_extract", cols_tx,
                                pd.read_sql(sql_tx, conn, params={"nm": name}, chunksize=chunk_size)
                            )
                        finally:
                            tx_sema.release()
                    # else: 无姓名则不建表

                    # 干掉默认空 sheet
                    if "Sheet" in wb.sheetnames:
                        del wb["Sheet"]
                    wb.save(path)

                dt = time.time() - t0
                return True, wb_id, path, f"ok in {dt:.2f}s"
            except Exception as e:
                if attempt < 4:
                    jitter_sleep(0.35, 1.8, attempt)
                    continue
                return False, wb_id, path, f"failed after retries: {repr(e)}"

    # === 并发执行 ===
    start = time.time()
    ok = 0
    err = 0
    print_every = max(1, args.print_every)
    with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="WB") as ex:
        futures = [ex.submit(process_one, r) for _, r in projects.iterrows()]
        for _, fut in enumerate(as_completed(futures), 1):
            success, wb_id, path, msg = fut.result()
            if success:
                ok += 1
                if (ok % print_every == 0) or (ok == total):
                    print(f"[OK] {ok}/{total} → {path} ({msg})")
            else:
                err += 1
                print(f"[ERR] {wb_id} → {path} ({msg})")

    elapsed = time.time() - start
    print(f"[DONE] 成功 {ok}，失败 {err}，总耗时 {elapsed:.1f}s。输出：{out_dir}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[ABORT] 用户中断。")
