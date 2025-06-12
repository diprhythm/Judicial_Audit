import os
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# —————— 配置区 ——————
HOST            = 'localhost'
PORT            = 3306
USER            = 'root'
PASSWORD        = '010203'

DB_MAIN         = '三平台合并数据'
DB_BANK         = '银行流水'
DB_OTHER        = '其它文件'

OUTPUT_DIR      = r'./output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

MAX_WORKERS     = 4         # 并发线程数，可按 CPU 核心数调整
CHUNK_SIZE_MAIN = 100_000   # 主表分块大小
CHUNK_SIZE_BANK = 100_000   # 银行流水分块大小

# —————— 日志配置 ——————
logging.basicConfig(
    level    = logging.INFO,
    format   = '%(asctime)s [%(levelname)s] %(message)s',
    datefmt  = '%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# —————— 建立数据库连接 ——————
engine_main  = create_engine(
    f'mysql+pymysql://{USER}:{PASSWORD}@{HOST}:{PORT}/{DB_MAIN}?charset=utf8mb4'
)
engine_bank  = create_engine(
    f'mysql+pymysql://{USER}:{PASSWORD}@{HOST}:{PORT}/{DB_BANK}?charset=utf8mb4'
)
engine_other = create_engine(
    f'mysql+pymysql://{USER}:{PASSWORD}@{HOST}:{PORT}/{DB_OTHER}?charset=utf8mb4'
)

# —————— 读取服务商清单 ——————
logger.info("载入“落地服务商清单”…")
df_landing = pd.read_sql(
    'SELECT 序号, 现用名_match, 曾用名_match FROM `落地服务商清单`',
    engine_other
).dropna(subset=['现用名_match'])
logger.info("共 %d 个服务商记录。", len(df_landing))

def stream_query_to_sheet(sql: str, engine, params: tuple,
                          sheet_name: str, wb: Workbook, chunk_size: int) -> int:
    ws = wb.create_sheet(sheet_name)
    first = True
    total = 0
    for chunk in pd.read_sql_query(sql, engine, params=params, chunksize=chunk_size):
        if first:
            for row in dataframe_to_rows(chunk, index=False, header=True):
                ws.append(row)
            first = False
        else:
            for row in dataframe_to_rows(chunk, index=False, header=False):
                ws.append(row)
        total += len(chunk)
    logger.info("    [%s] 共写入 %d 行", sheet_name, total)
    return total

def process_provider(row):
    seq      = row['序号']
    cur_name = row['现用名_match']
    old_name = row['曾用名_match']
    aliases  = [cur_name]
    if pd.notna(old_name) and old_name != cur_name:
        aliases.append(old_name)

    ph_alias = ','.join(['%s'] * len(aliases))
    sql_main = (
        f"SELECT * FROM `002_下发汇总表_20250411_剔除非30_清洗非客户公司` "
        f"WHERE 服务公司名称_match IN ({ph_alias})"
    )

    logger.info("开始处理 → %s（别名: %s）", cur_name, aliases)
    wb = Workbook(write_only=True)

    # ——— 1. 服务商下发数据 & 子服务商过滤收集 ———
    acct_totals, acct_subservs, subserv_set = {}, {}, set()
    ws1 = wb.create_sheet('1.服务商下发数据')
    first, total1 = True, 0
    for chunk in pd.read_sql_query(sql_main, engine_main,
                                   params=tuple(aliases),
                                   chunksize=CHUNK_SIZE_MAIN):
        if first:
            for r in dataframe_to_rows(chunk, index=False, header=True):
                ws1.append(r)
            first = False
        else:
            for r in dataframe_to_rows(chunk, index=False, header=False):
                ws1.append(r)
        total1 += len(chunk)
        # 新：只把既非 NaN 且 strip() 后非空字符串的子服务商加入 set
        for acct, sub in zip(chunk['账户名称_match'], chunk['子服务商公司名_match']):
            if pd.isna(acct):
                continue
            acct_totals[acct] = acct_totals.get(acct, 0) + 1
            if pd.notna(sub) and str(sub).strip():
                clean_sub = str(sub).strip()
                acct_subservs[acct] = acct_subservs.get(acct, 0) + 1
                subserv_set.add(clean_sub)
    logger.info("    [1.服务商下发数据] 写入 %d 行", total1)

    # ——— 子服务商过滤：剔除“仅出现在子服务商行”的账户 ———
    accounts_keep = [
        acct for acct, tot in acct_totals.items()
        if acct_subservs.get(acct, 0) < tot
    ]

    # ——— 2. 富民银行下发 ———
    sheet2_count = 0
    if accounts_keep:
        ph_acct = ','.join(['%s'] * len(accounts_keep))
        sql_fumin = (
            f"SELECT * FROM `富民银行流水` "
            f"WHERE 账户名_match IN ({ph_alias}) "
            f"  AND 对手户名_match IN ({ph_acct})"
        )
        sheet2_count = stream_query_to_sheet(
            sql_fumin, engine_bank,
            tuple(aliases) + tuple(accounts_keep),
            '2.富民银行下发', wb, CHUNK_SIZE_BANK
        )
    else:
        logger.info("    [2.富民银行下发] 无符合条件账户，跳过")

    # ——— 3. 其它银行下发 ———
    sheet3_count = 0
    if accounts_keep:
        sql_other = (
            f"SELECT * FROM `其它银行流水` "
            f"WHERE 账户名_match IN ({ph_alias}) "
            f"  AND 对手户名_match IN ({ph_acct})"
        )
        sheet3_count = stream_query_to_sheet(
            sql_other, engine_bank,
            tuple(aliases) + tuple(accounts_keep),
            '3.其它银行下发', wb, CHUNK_SIZE_BANK
        )
    else:
        logger.info("    [3.其它银行下发] 无符合条件账户，跳过")

    # ——— 4/5. 子服务商流水 ———
    if subserv_set:
        sub_list = list(subserv_set)
        ph_sub   = ','.join(['%s'] * len(sub_list))
        all_accts = list(acct_totals.keys())
        ph_all    = ','.join(['%s'] * len(all_accts))

        sql_sub_fumin = (
            f"SELECT * FROM `富民银行流水` "
            f"WHERE 账户名_match IN ({ph_sub}) "
            f"  AND 对手户名_match IN ({ph_all})"
        )
        stream_query_to_sheet(
            sql_sub_fumin, engine_bank,
            tuple(sub_list) + tuple(all_accts),
            '4.子服务商富民银行下发', wb, CHUNK_SIZE_BANK
        )

        sql_sub_other = (
            f"SELECT * FROM `其它银行流水` "
            f"WHERE 账户名_match IN ({ph_sub}) "
            f"  AND 对手户名_match IN ({ph_all})"
        )
        stream_query_to_sheet(
            sql_sub_other, engine_bank,
            tuple(sub_list) + tuple(all_accts),
            '5.子服务商其它银行下发', wb, CHUNK_SIZE_BANK
        )
        has_sub = '存在子服务商'
    else:
        logger.info("    无子服务商，跳过 4/5 两个表")
        has_sub = '无子服务商'

    # ——— 命名：判断流水缺失 ———
    if sheet2_count == 0 and sheet3_count == 0:
        flow_status = '未匹配到任何流水'
    elif sheet2_count == 0:
        flow_status = '无富民银行下发'
    elif sheet3_count == 0:
        flow_status = '无其它银行下发'
    else:
        flow_status = '正常'

    display_name = f"{cur_name}（{old_name}）" if pd.notna(old_name) and old_name != cur_name else cur_name
    filename = f"{seq}、{display_name}_{has_sub}_{flow_status}.xlsx"
    path     = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    logger.info("完成 ← %s，文件已保存：%s", cur_name, filename)

def main():
    providers = df_landing.to_dict('records')
    logger.info("共 %d 个服务商 → 启动 %d 线程并行处理", len(providers), MAX_WORKERS)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as exe:
        futures = [exe.submit(process_provider, row) for row in providers]
        for fut in as_completed(futures):
            if fut.exception():
                logger.error("处理异常：%s", fut.exception())

if __name__ == '__main__':
    main()
