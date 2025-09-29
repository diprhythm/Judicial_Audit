# coding: utf-8
r"""
登记信息生成_optimize
- 从统计表和流水文件夹生成《登记信息.xlsx》，并将文件按规则归档：
  1已登记 / 2被跳过 / 3重复账号
- 账户信息工作表筛选：选择含“账户信息”并排除含“关联子账户信息”的表
- 并发安全：查→分→写映射 在同一把锁内（原子），搬运放锁外
- 线程数：基于CPU与文件数自动决定，最大不超过 61（Fuck Microsoft）
- .xls 兼容：需要 xlrd==1.2.0；否则提示转 .xlsx

依赖：pandas, openpyxl（必要），xlrd==1.2.0（若需读取 .xls）
"""

import os
import re
import sys
import shutil
import logging
import warnings
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style",
    category=UserWarning,
)

import pandas as pd
from openpyxl import load_workbook, Workbook
import numpy as np


# =========================
# 常量
# =========================
HEADER_ROW = 5
STATISTICS_SHEET = '统计表'
OUTPUT_FILE_NAME = '登记信息.xlsx'
SUBDIR_REGISTERED = '1已登记'
SUBDIR_SKIPPED = '2被跳过'
SUBDIR_DUPLICATE = '3重复账号'
WIN_THREAD_CAP = 61

ALLOW_RE = re.compile(r'^\s*账户信息(\d+)?\s*$')
DENY_RE  = re.compile(r'关联子账户信息')

SAFE_NAME_RE = re.compile(r'[<>:"/\\|?*\x00-\x1F]')  # Windows 不允许的字符


# =========================
# 日志
# =========================
def setup_logger(folder: Path) -> logging.Logger:
    log_path = folder / 'run_log.txt'
    logger = logging.getLogger('register_runner')
    logger.setLevel(logging.INFO)
    logger.propagate = False  # 防止重复输出到根logger

    if not logger.handlers:
        fmt = logging.Formatter('%(asctime)s %(levelname)s: %(message)s')

        # 文件 handler
        fh = logging.FileHandler(str(log_path), encoding='utf-8')
        fh.setFormatter(fmt)
        logger.addHandler(fh)

        # 控制台 handler
        sh = logging.StreamHandler(sys.stdout)
        sh.setFormatter(fmt)
        logger.addHandler(sh)

    return logger


# =========================
# 工具函数
# =========================
def strip_trailing_dot_zero(s: str) -> str:
    """只移除末尾恰好是'.0'的情况；不吞尾零。"""
    if s is None:
        return None
    return re.sub(r'\.0$', '', str(s).strip())


def normalize_number(val) -> str | None:
    """把单元格值安全转成字符串（仅去掉末尾'.0'），保留所有有效尾零。"""
    if val is None:
        return None
    s = strip_trailing_dot_zero(str(val))
    s = s.strip()
    if s.lower() in ('nan', 'nat'):
        return None
    return s or None


def digits_only(s: str | None) -> str:
    """提取字符串中的数字序列；None -> ''。"""
    if not s:
        return ''
    return re.sub(r'\D', '', str(s))


def digit_match(a: str | None, b: str | None, min_len: int = 12) -> bool:
    """
    “数字匹配”：仅比较两侧的数字序列。
    若较短者的数字串是较长者数字串的子串，且较短者长度 >= min_len，则认为匹配。
    """
    da, db = digits_only(a), digits_only(b)
    if not da or not db:
        return False
    if da == db:
        return True
    sa, sb = (da, db) if len(da) <= len(db) else (db, da)
    if len(sa) < min_len:
        return False
    return sa in sb


def format_date(val) -> str | None:
    """格式化日期为 YYYY.MM.DD；无法解析则返回 None。"""
    if val is None:
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.strftime('%Y.%m.%d')
    s = str(val).strip()
    # 8位数字串：YYYYMMDD
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}.{s[4:6]}.{s[6:]}"
    dt = pd.to_datetime(s, errors='coerce')
    if pd.notna(dt):
        return dt.strftime('%Y.%m.%d')
    return None


def to_datetime_or_na(val):
    """尽力把值转为 pandas.Timestamp；失败返回 NaT。"""
    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.to_datetime(val, errors='coerce')
    s = str(val).strip()
    if not s:
        return pd.NaT
    if len(s) == 8 and s.isdigit():
        try:
            return pd.to_datetime(f"{s[:4]}-{s[4:6]}-{s[6:]}", errors='coerce')
        except Exception:
            return pd.NaT
    return pd.to_datetime(s, errors='coerce')


def clean_input_path(path: str) -> str:
    """去除路径中的不可见字符（BOM/方向控制符等）"""
    bad_chars = '\ufeff\u202a\u202b\u202c\u200e\u200f'
    return path.strip().translate({ord(c): None for c in bad_chars})


def sanitize_filename(name: str) -> str:
    """替换Windows不允许字符；去掉末尾空格和点。"""
    name = SAFE_NAME_RE.sub('_', name)
    name = name.rstrip(' .')
    return name


def decide_thread_count(num_files: int) -> int:
    """根据 CPU 与文件数决定线程数；不超过 61。"""
    cpus = os.cpu_count() or 4
    if num_files <= 1:
        return 1
    # I/O 密集，允许开到 2x CPU，但不超文件数/61
    base = min(num_files, cpus * 2)
    return max(2, min(WIN_THREAD_CAP, base, num_files))


# =========================
# Excel 读取工具
# =========================
def pick_account_info_sheetnames(path: Path) -> list[str]:
    """列出所有允许的“账户信息*”sheet（排除“关联子账户信息*”）。"""
    # 用 pandas.ExcelFile 拿 sheet_names（支持 xlsx/xls）
    engine = None
    suf = path.suffix.lower()
    if suf in ('.xlsx', '.xlsm'):
        engine = 'openpyxl'
    elif suf == '.xls':
        engine = 'xlrd'  # 需要 xlrd==1.2.0
    try:
        with pd.ExcelFile(path, engine=engine) as xf:
            names = xf.sheet_names
    except Exception as e:
        raise RuntimeError(f'无法读取工作簿以获取工作表列表：{path.name}。若为 .xls，请安装 xlrd==1.2.0 或转换为 .xlsx') from e

    # 排除关联子账户信息；允许严格 ^账户信息(\d+)?$
    out = []
    for sn in names:
        if DENY_RE.search(sn):
            continue
        if ALLOW_RE.match(sn):
            out.append(sn)
    return out


def read_excel_safe(path: Path, sheet_name, dtype=str) -> pd.DataFrame:
    """按扩展名选择引擎读取 Excel；.xls 需要 xlrd==1.2.0。"""
    suf = path.suffix.lower()
    engine = None
    if suf in ('.xlsx', '.xlsm'):
        engine = 'openpyxl'
    elif suf == '.xls':
        engine = 'xlrd'
    return pd.read_excel(path, sheet_name=sheet_name, dtype=dtype, engine=engine)


def read_account_info_df(path: Path, logger: logging.Logger) -> pd.DataFrame:
    """读取并合并所有有效的 '账户信息*' 表；若无则返回空表。"""
    sheets = pick_account_info_sheetnames(path)
    if not sheets:
        return pd.DataFrame()
    dfs = []
    for sn in sheets:
        try:
            df = read_excel_safe(path, sn, dtype=str)
            if isinstance(df, pd.DataFrame) and not df.empty:
                dfs.append(df)
        except Exception:
            logger.warning("读取账户信息工作表失败: %s | %s", sn, path.name)
            continue
    if not dfs:
        return pd.DataFrame()
    # 统一去掉末尾 .0、清理空白
    df = pd.concat(dfs, ignore_index=True)
    for col in ('交易账号', '交易卡号', '账号开户银行', '开户网点', '账号开户时间', '销户日期'):
        if col in df.columns:
            df[col] = df[col].map(normalize_number)
    return df


def read_unique_values(df: pd.DataFrame, column: str) -> list[str]:
    """取指定列的去重非空字符串值，安全移除末尾 '.0'。"""
    if column not in df.columns:
        return []
    ser = df[column].dropna().astype(str).map(strip_trailing_dot_zero)
    ser = ser.map(lambda x: x.strip()).replace({'nan': ''}).replace({'NaT': ''})
    ser = ser[ser != '']
    return list(pd.unique(ser))


# =========================
# 统计表读取
# =========================
def read_statistics_info(path: Path, logger: logging.Logger):
    """读取统计表中的已用序号、账号->序号、卡号->序号映射。"""
    wb = load_workbook(path, data_only=True)
    sheet = wb[STATISTICS_SHEET] if STATISTICS_SHEET in wb.sheetnames else wb.active

    headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[HEADER_ROW]]
    header_map = {}
    for idx, h in enumerate(headers, start=1):
        header_map[h] = idx
        header_map[h.replace('\n', '')] = idx  # 容错：去掉换行

    def col(name: str):
        return header_map.get(name)

    used_serials = set()
    account_serial = {}
    card_serial = {}

    for row in sheet.iter_rows(min_row=HEADER_ROW + 1):
        # 序号
        serial_cell = row[col('已取序号手动添加') - 1] if col('已取序号手动添加') else None
        serial_value = serial_cell.value if serial_cell else None
        serial = None
        if serial_value is not None:
            try:
                serial = int(str(serial_value).split('.')[0])
                if serial > 0:
                    used_serials.add(serial)
            except Exception:
                pass

        # 账号/卡号
        acc_val = row[col('账号') - 1].value if col('账号') else None
        card_val = row[col('卡号') - 1].value if col('卡号') else None
        acc_str = normalize_number(acc_val)
        card_str = normalize_number(card_val)

        if serial is not None:
            if acc_str:
                account_serial.setdefault(acc_str, serial)
            if card_str:
                card_serial.setdefault(card_str, serial)

    logger.info("统计表载入完成：已用序号 %d 个，账号映射 %d 条，卡号映射 %d 条",
                len(used_serials), len(account_serial), len(card_serial))
    return used_serials, account_serial, card_serial


# =========================
# 序号与映射（注意：外层统一持锁）
# =========================
def next_serial(used: set[int]) -> int:
    """分配下一个可用正整数序号（找最小可用）。"""
    i = 1
    while i in used:
        i += 1
    used.add(i)
    return i


def find_existing_serial(account: str | None, cards: list[str],
                         account_map: dict, card_map: dict) -> int | None:
    """在映射表中查找已存在的序号（account 或任何 card）。"""
    if account:
        serial = account_map.get(account) or card_map.get(account)
        if serial is not None:
            return serial
    for c in cards:
        serial = card_map.get(c) or account_map.get(c)
        if serial is not None:
            return serial
    return None


# =========================
# 文件搬运与输出工作簿
# =========================
def ensure_dirs(base: Path):
    for name in (SUBDIR_REGISTERED, SUBDIR_SKIPPED, SUBDIR_DUPLICATE):
        (base / name).mkdir(parents=True, exist_ok=True)


def create_output_book(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.append([
        '文件名', '已取序号', '是否整理', '币种', '账户名', '开户行',
        '账号', '卡号', '是否取得', '取得时间', '*详细说明',
        '*是否需重取', '⑤编制人', '已取得交易期间'
    ])
    return wb


def rename_and_move(src: Path, dest_dir: Path, serial, acquire_time: str, logger: logging.Logger) -> str:
    dest_dir.mkdir(parents=True, exist_ok=True)
    serial_str = serial if isinstance(serial, str) else f"{int(serial):04d}"
    base, ext = src.stem, src.suffix
    name = f"{serial_str}、{base}({acquire_time}){ext}"
    name = sanitize_filename(name)
    target = dest_dir / name
    # 若同名存在，添加去重后缀
    if target.exists():
        k = 2
        while True:
            cand = dest_dir / sanitize_filename(f"{serial_str}、{base}({acquire_time})_{k}{ext}")
            if not cand.exists():
                target = cand
                break
            k += 1
    shutil.move(str(src), str(target))
    logger.info("移动文件：%s -> %s", src.name, target.name)
    return target.name


# =========================
# 单文件处理
# =========================
def process_file(file_path: Path,
                 stats_accounts: dict, stats_cards: dict, used_serials: set[int],
                 acquire_time: str, author: str,
                 lock: threading.Lock,
                 folders: tuple[Path, Path, Path],
                 logger: logging.Logger,
                 min_len_digits: int = 12) -> dict | None:
    registered_dir, skipped_dir, duplicate_dir = folders
    basename = file_path.name
    logger.info("开始处理：%s", basename)

    # 1) 读“提取”
    try:
        tiqu = read_excel_safe(file_path, '提取', dtype=str)
    except Exception:
        logger.exception("读取【提取】失败：%s", basename)
        # 移动到 2被跳过
        try:
            shutil.move(str(file_path), str(skipped_dir / basename))
        except Exception:
            pass
        return {'文件名': basename, '*详细说明': '无法读取提取表'}

    # 2) 读“整理表”（可选，只为币种）
    try:
        zhengli = read_excel_safe(file_path, '整理表', dtype=str)
    except Exception:
        logger.warning("整理表读取失败：%s", basename)
        zhengli = pd.DataFrame()

    # 3) 读“账户信息*”（正则+排除）
    account_info = read_account_info_df(file_path, logger)

    record = {'文件名': basename}

    # 账户名
    names = read_unique_values(tiqu, '本账号名称')
    if not names:
        record.update({'账户名': '无账户名'})
        try:
            shutil.move(str(file_path), str(skipped_dir / basename))
        except Exception:
            pass
        logger.warning("跳过（无账户名）：%s", basename)
        return record
    record['账户名'] = ';'.join(names)

    # 账号
    accounts = read_unique_values(tiqu, '本账号')
    if len(accounts) == 1:
        account = accounts[0]
        record['账号'] = account
    else:
        record['账号'] = '非唯一账号 请检查'
        account = None

    # 卡号
    cards = read_unique_values(tiqu, '本卡号')
    record['卡号'] = '、'.join(cards)

    # 币种（来自整理表）
    currency = read_unique_values(zhengli, '交易币种') if not zhengli.empty else []
    record['币种'] = ';'.join(currency)

    # 开户行 + 开户/销户时间（账户信息*）
    bank = '未匹配成功'
    open_time_fmt = None
    close_time_fmt = None

    if account and not account_info.empty:
        # 第一层：精确匹配
        acc_series = account_info.get('交易账号', pd.Series(dtype=str)).map(normalize_number)
        card_series = account_info.get('交易卡号', pd.Series(dtype=str)).map(normalize_number)
        exact_mask = (acc_series == account) | (card_series == account)
        subset = account_info[exact_mask]

        # 第二层：数字匹配（仅当精确匹配为空）
        if subset.empty:
            # 对每一行做 digit_match，命中任一列即算
            matches = []
            for idx, row in account_info.iterrows():
                a = row.get('交易账号')
                c = row.get('交易卡号')
                if digit_match(account, a, min_len_digits) or digit_match(account, c, min_len_digits):
                    matches.append(idx)
            if matches:
                subset = account_info.loc[matches]

        if not subset.empty:
            # 开户行：取第一条命中的银行与网点，直接拼接（无分隔、无空格）
            ob = None
            br = None
            for _, r in subset.iterrows():
                if ob is None:
                    val = r.get('账号开户银行')
                    if isinstance(val, str) and val.strip():
                        ob = val.strip()
                if br is None:
                    val = r.get('开户网点')
                    if isinstance(val, str) and val.strip():
                        br = val.strip()
                if ob is not None and br is not None:
                    break
            bank = f"{ob or ''}{br or ''}" if (ob or br) else '无数据'

            # 开户/销户时间：最早开户、最晚销户
            open_times = subset.get('账号开户时间')
            close_times = subset.get('销户日期')
            if open_times is not None:
                ots = pd.to_datetime(open_times.map(to_datetime_or_na), errors='coerce')
                if ots.notna().any():
                    open_time_fmt = ots.min().strftime('%Y.%m.%d')
            if close_times is not None:
                cts = pd.to_datetime(close_times.map(to_datetime_or_na), errors='coerce')
                if cts.notna().any():
                    close_time_fmt = cts.max().strftime('%Y.%m.%d')

    record['开户行'] = bank

    # 详细说明：开户/销户 + 余额差异
    detail_parts = []
    if open_time_fmt:
        detail_parts.append(f"开户时间{open_time_fmt}")
    else:
        detail_parts.append("开户时间匹配失败")
    if close_time_fmt:
        detail_parts.append(f"销户时间{close_time_fmt}")

    diff_desc = None
    if '公式校验' in tiqu.columns and '日期' in tiqu.columns:
        check_series = pd.to_numeric(tiqu['公式校验'], errors='coerce')
        dates_series = pd.to_datetime(tiqu['日期'], errors='coerce')
        valid = check_series.notna() & dates_series.notna()
        if valid.any():
            chk = check_series[valid]
            dts = dates_series[valid]
            # 非“近似零”的视为差异
            nonzero_mask = ~np.isclose(chk.values.astype(float), 0.0, atol=0.005)
            if nonzero_mask.any():
                nz_idx = chk.index[nonzero_mask]
                first_date = dts.loc[nz_idx[0]].strftime('%Y年%m月')
                last_date = dts.loc[nz_idx[-1]].strftime('%Y年%m月')
                # 最大绝对差异
                max_idx = (chk.abs()).idxmax()
                max_val = chk.loc[max_idx]
                max_val_str = (
                    str(int(max_val)) if pd.notna(max_val) and float(max_val).is_integer()
                    else f"{float(max_val):.2f}"
                )
                diff_desc = f"{first_date}至{last_date}存在余额差异 余额差异最大为{max_val_str}"
    if diff_desc:
        detail_parts.append(diff_desc)

    record['*详细说明'] = f"{acquire_time}取得：" + '；'.join(detail_parts)

    # 交易期间
    if '日期' in tiqu.columns and not tiqu['日期'].dropna().empty:
        dts = pd.to_datetime(tiqu['日期'].dropna(), errors='coerce').dropna()
        if not dts.empty:
            start = dts.min().strftime('%Y.%m.%d')
            end = dts.max().strftime('%Y.%m.%d')
            record['已取得交易期间'] = f'{start}-{end}'
        else:
            record['已取得交易期间'] = ''
    else:
        record['已取得交易期间'] = ''

    # 固定标志
    record['是否整理'] = '是'
    record['是否取得'] = '是'
    record['取得时间'] = acquire_time
    record['*是否需重取'] = '否'
    record['⑤编制人'] = author

    # ===== 原子区：查→分→写映射（锁内） =====
    assigned_serial = None
    is_duplicate = False
    with lock:
        existing_serial = find_existing_serial(account, cards, stats_accounts, stats_cards)
        if existing_serial is not None:
            assigned_serial = existing_serial
            is_duplicate = True
        else:
            assigned_serial = next_serial(used_serials)
            if account:
                stats_accounts[account] = assigned_serial
            for c in cards:
                stats_cards[c] = assigned_serial

    # 记录序号
    record['已取序号'] = f'已取序号{assigned_serial}' if is_duplicate else assigned_serial

    # ===== 锁外：重命名/搬运（I/O）=====
    dest = duplicate_dir if is_duplicate else registered_dir
    try:
        final_name = rename_and_move(file_path, dest, assigned_serial, acquire_time, logger)
        record['文件名'] = final_name
    except Exception:
        # 搬运失败，仍返回记录
        pass

    logger.info("%s 完成：%s | 序号=%s | 重复=%s", basename, record.get('文件名', basename),
                assigned_serial, is_duplicate)
    return record


# =========================
# 主流程
# =========================
def main():
    # 输入
    stats_path_str = clean_input_path(input('请输入统计表路径: '))
    folder_str = input('请输入拆分后流水的存放路径: ').strip()
    acquire_time = input('请输入取得时间: ').strip()
    author = input('请输入编制人: ').strip()

    folder = Path(folder_str).resolve()
    stats_path = Path(stats_path_str).resolve()

    logger = setup_logger(folder)
    logger.info("=== 任务开始 ===")

    # 目录与输出
    ensure_dirs(folder)
    output_path = folder / OUTPUT_FILE_NAME
    wb = create_output_book(output_path)
    ws = wb.active

    # 读统计表
    used_serials, stats_accounts, stats_cards = read_statistics_info(stats_path, logger)

    # 待处理文件列表
    excluded_names = {
        OUTPUT_FILE_NAME,
        stats_path.name if stats_path.parent.resolve() == folder else None,
    }
    excluded_names.discard(None)
    excluded_dirs = {SUBDIR_REGISTERED, SUBDIR_SKIPPED, SUBDIR_DUPLICATE}

    files = []
    for name in os.listdir(folder):
        p = folder / name
        if p.is_dir():
            if name in excluded_dirs:
                continue
            else:
                continue  # 不遍历子目录
        if name in excluded_names:
            continue
        if name.startswith('~$'):
            continue
        if name.lower().endswith(('.xls', '.xlsx', '.xlsm')):
            files.append(p)

    if not files:
        print("没有找到可处理的 Excel 文件。")
        logger.info("没有找到可处理的 Excel 文件。")
        return

    # 线程数
    thread_count = decide_thread_count(len(files))
    logger.info("文件数=%d | 线程数=%d", len(files), thread_count)

    # 处理
    lock = threading.Lock()
    registered = folder / SUBDIR_REGISTERED
    skipped = folder / SUBDIR_SKIPPED
    duplicated = folder / SUBDIR_DUPLICATE
    folders = (registered, skipped, duplicated)

    results = []
    stats = {'ok': 0, 'dup': 0, 'skip': 0}
    total = len(files)
    done = 0
    t0 = time.monotonic()

    def log_progress():
        elapsed = time.monotonic() - t0
        avg = elapsed / done if done else 0.0
        remain = avg * (total - done) if done else 0.0
        def fmt(s):
            m, s = divmod(int(s), 60)
            return f"{m:02d}:{s:02d}"
        logger.info("进度: %d/%d | 用时 %s | 预计剩余 %s", done, total, fmt(elapsed), fmt(remain))

    if thread_count == 1:
        for f in files:
            rec = process_file(f, stats_accounts, stats_cards, used_serials,
                               acquire_time, author, lock, folders, logger)
            done += 1
            log_progress()
            if rec:
                results.append(rec)
    else:
        with ThreadPoolExecutor(max_workers=thread_count) as executor:
            futs = [executor.submit(process_file, f, stats_accounts, stats_cards, used_serials,
                                    acquire_time, author, lock, folders, logger)
                    for f in files]
            for fut in as_completed(futs):
                try:
                    rec = fut.result()
                except Exception as e:
                    logger.exception("处理线程异常：%s", e)
                    rec = None
                done += 1
                log_progress()
                if rec:
                    results.append(rec)

    # （可选）按序号排序写出，便于阅读
    def extract_serial(v):
        if isinstance(v, int):
            return v
        if isinstance(v, str) and v.startswith('已取序号'):
            try:
                return int(v.replace('已取序号', '').strip())
            except:
                return 10**9
        return 10**9
    results.sort(key=lambda r: extract_serial(r.get('已取序号', '')))

    # 汇总写出（原子替换，避免占用损坏）
    for rec in results:
        ws.append([
            rec.get('文件名', ''),
            rec.get('已取序号', ''),
            rec.get('是否整理', ''),
            rec.get('币种', ''),
            rec.get('账户名', ''),
            rec.get('开户行', ''),
            rec.get('账号', ''),
            rec.get('卡号', ''),
            rec.get('是否取得', ''),
            rec.get('取得时间', ''),
            rec.get('*详细说明', ''),
            rec.get('*是否需重取', ''),
            rec.get('⑤编制人', ''),
            rec.get('已取得交易期间', ''),
        ])
        # 统计
        v = rec.get('已取序号', '')
        if isinstance(v, str) and v.startswith('已取序号'):
            stats['dup'] += 1
        elif isinstance(v, int):
            stats['ok'] += 1
        if rec.get('*详细说明', '') == '无法读取提取表' or rec.get('账户名', '') == '无账户名':
            stats['skip'] += 1

    tmp_path = output_path.with_suffix('.tmp.xlsx')
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, output_path)
    except PermissionError:
        logger.error("保存失败：目标文件可能正被 Excel 打开 -> %s", output_path)
        print("保存失败：请关闭正在打开的《登记信息.xlsx》后重试。")
        return

    logger.info("登记信息已生成: %s", output_path)
    logger.info("统计：登记=%d | 重复=%d | 跳过=%d", stats['ok'], stats['dup'], stats['skip'])
    print(f'登记信息已生成: {output_path}')
    print(f'统计：登记={stats["ok"]} | 重复={stats["dup"]} | 跳过={stats["skip"]}')
    logger.info("=== 任务结束 ===")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"运行失败：{e}", file=sys.stderr)
        raise
